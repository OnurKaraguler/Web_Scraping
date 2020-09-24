from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait as WDW
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select        # for dropdown
import time
import openpyxl as O

exec_path = r"C:\Users\USER\Documents\Selenium\chromedriver.exe"
URL = r"http://inderpsingh.blogspot.com/2014/08/demowebapp_24.html"
Excel_file = "C:\\Python_Examples\\Web_Scraping\\Selenium\\Selenium Python tutorial\\excel\\JourneyPlanner.xlsx"
Excel_worksheet = "Sayfa1"
distance_id_locator = "distance"
distance = 0
speed_id_locator = "speed"
speed = 45
time_id_locator = "hours"       # dropdown menu
calculate_css_locator = "#Blog1 > div.blog-posts.hfeed > div > div > div.post-outer > div.post.hentry > div.post-body.entry-content > div:nth-child(1) > div > form > button:nth-child(20)"      # calculate button, right click-copy-CSS Selector
result_id_locator = "result"
message_id_locator = "message"
wait_time_out = 5
driver = webdriver.Chrome(executable_path=exec_path)
driver.get(URL)
wait_variable = WDW(driver,wait_time_out)
driver.maximize_window()

driver.execute_script("window.scrollBy(0,240)", "")  # Ekranı aşağı kaydırıyor
distance_element = wait_variable.until(EC.presence_of_element_located((By.ID,distance_id_locator)))
speed_element = wait_variable.until(EC.presence_of_element_located((By.ID,speed_id_locator)))
time_element = Select(wait_variable.until(EC.presence_of_element_located((By.ID,time_id_locator))))
calculate_element = wait_variable.until(EC.presence_of_element_located((By.CSS_SELECTOR,calculate_css_locator)))
result_element = wait_variable.until(EC.presence_of_element_located((By.ID,result_id_locator)))
message_element = wait_variable.until(EC.presence_of_element_located((By.ID, message_id_locator)))
wb = O.load_workbook(Excel_file)
ws = wb[Excel_worksheet]
for r in range(2,ws.max_row + 1):
    d = str(ws.cell(r,1).value)
    distance_element.clear()
    distance_element.send_keys(d)
    s = str(ws.cell(r,2).value)
    speed_element.clear()
    speed_element.send_keys(s)
    t = str(ws.cell(r,3).value)
    time_element.select_by_visible_text(t)
    calculate_element.click()
    time.sleep(1)
    e = str(ws.cell(r,4).value)
    if str(e) in result_element.text or str(e) in message_element.text:
        ws.cell(r,5).value = "Pass"
    else:
        ws.cell(r,5).value = "Fail"
    wb.save(Excel_file)         # Önemli
    wb.close()          # Önemli

# row_num = ws.max_row
# col_num = ws.max_column