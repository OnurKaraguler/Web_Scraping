# pip install openpyxl
import openpyxl as O

Excel_file = "C:\\Python_Examples\\Web_Scraping\\Selenium\\Selenium Python tutorial\\excel\\JourneyPlanner.xlsx"
Excel_worksheet = "Sayfa1"
wb = O.load_workbook(Excel_file)
ws = wb[Excel_worksheet]
row_num = ws.max_row
col_num = ws.max_column
print("The no. of rows is", row_num, "and the number of columns is", col_num)
row = 2
print("distance = ", ws.cell(row, 1).value)
print("speed = ", ws.cell(row, 2).value)
print("Travel hours/day = ", ws.cell(row, 3).value)
print("Expected Result = ", ws.cell(row, 4).value)