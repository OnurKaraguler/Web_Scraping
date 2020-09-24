from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait as WDW
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select        # for dropdown
import openpyxl as O
from Web_Scraping.Selenium.userInfo import username,password
import time
import pyautogui as P

class eksisozluk:
    def __init__(self,username,password):

        self.username = username
        self.password = password

    def signIn(self):
        exec_path = r"C:\Users\USER\Documents\Selenium\chromedriver.exe"         # r text yapar
        URL = r"https://eksisozluk.com/giris?returnUrl=https%3A%2F%2Feksisozluk.com%2F"
        Excel_file = "C:\\Python_Examples\\Web_Scraping\\Selenium\\excel\\EksisozlukGundem.xlsx"
        Excel_worksheet = "Sayfa1"
        wb = O.load_workbook(Excel_file)
        ws = wb[Excel_worksheet]
        gundem_locator = "//*[@id='partial-index']/ul"
        gundem_menu_loc = "//*[@id='quick-index-nav']/li[2]/a"
        wait_time_out = 5
        self.browser = webdriver.Chrome(executable_path=exec_path)
        self.browser.get(URL)
        time.sleep(1)
        wait_variable = WDW(self.browser, wait_time_out)
        self.browser.maximize_window()

        username = self.browser.find_element_by_xpath("//*[@id='username']").send_keys(self.username)
        password = self.browser.find_element_by_xpath("//*[@id='password']").send_keys(self.password)
        time.sleep(1)
        self.browser.find_element_by_xpath("//*[@id='login-form-container']/form/fieldset/div[4]/button").click()
        time.sleep(1)

        # P.moveTo(200,220,2)
        # P.leftClick()
        # time.sleep(1)

        gundem_menu_element = wait_variable.until(EC.presence_of_element_located((By.XPATH, gundem_menu_loc)))
        gundem_menu_element.click()

        gundem_reyting_List = []
        linkList = []
        for i in range(1,31):
            try:
                gundemler = self.browser.find_element_by_xpath(f"//*[@id='partial-index']/ul/li[{i}]/a").text.split('\n')
                linkler = self.browser.find_element_by_xpath(f"//*[@id='partial-index']/ul/li[{i}]/a").get_attribute('href')
                gundem_reyting_List.append(gundemler)
                linkList.append(linkler)
            except:
                continue

        gundemList = []
        reytinglist = []
        for j in gundem_reyting_List:
            try:
                gundemList.append(j[0])
                reytinglist.append(j[1])
            except:
                reytinglist.append('0')
                continue

        sonList = list(zip(gundemList,reytinglist,linkList))
        print('-' * 20, 'GÜNDEM BAŞLIKLARI', '-' * 20, sep='\n')  # sep='\n' sırayla yapar
        for num,k in enumerate(sonList,1):
            gundem = k[0]
            reyting = k[1]
            link = k[2]
            print(f"{num}. Gündem: {gundem}\nReyting: {reyting}\nLink: {link} ")
            print('=' * 20)
            ws.cell(num+1, 1).value = gundem
            ws.cell(num+1, 2).value = int(reyting)
            ws.cell(num+1, 3).value = link

            wb.save(Excel_file)  # Önemli
            wb.close()  # Önemli

        self.browser.quit()

eksi = eksisozluk(username,password)
eksi.signIn()